"""
iPad & Stylus Distribution System — Streamlit
All features: Issue, Return, Dashboard, Records, Students (Add+Bulk), Reports
Deploy free: streamlit.io/cloud
"""

import streamlit as st
import sqlite3
import csv
import io
import os
import pandas as pd
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── PAGE CONFIG ────────────────────────────────────────────
st.set_page_config(
    page_title="iPad & Stylus Distribution",
    page_icon="📱",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ── CUSTOM CSS ─────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap');

html, body, [class*="css"] {
    font-family: 'Inter', sans-serif;
}

/* Sidebar styling */
[data-testid="stSidebar"] {
    background: #fff;
    border-right: 1px solid #e5e7eb;
}
[data-testid="stSidebar"] [data-testid="stMarkdownContainer"] p {
    font-size: 13px;
    color: #374151;
}

/* Main background */
.main { background: #f4f5f7; }
.block-container { padding: 1.5rem 2rem; }

/* Metric cards */
[data-testid="metric-container"] {
    background: #fff;
    border: 1px solid #e5e7eb;
    border-radius: 12px;
    padding: 16px 20px;
    box-shadow: 0 1px 3px rgba(0,0,0,.05);
}
[data-testid="metric-container"] label {
    font-size: 12px !important;
    color: #6b7280 !important;
    font-weight: 500 !important;
}
[data-testid="metric-container"] [data-testid="stMetricValue"] {
    font-size: 26px !important;
    font-weight: 700 !important;
    color: #111827 !important;
}

/* Buttons */
.stButton > button {
    background: #16a34a;
    color: white;
    border: none;
    border-radius: 8px;
    font-weight: 600;
    font-size: 13.5px;
    padding: 8px 20px;
    transition: all .15s;
    font-family: 'Inter', sans-serif;
}
.stButton > button:hover {
    background: #15803d;
    color: white;
}

/* Warning button style */
.warn-btn .stButton > button {
    background: #f59e0b;
}
.warn-btn .stButton > button:hover {
    background: #d97706;
}

/* Input fields */
.stTextInput > div > div > input,
.stSelectbox > div > div,
.stTextArea > div > div > textarea {
    border-radius: 8px !important;
    border: 1px solid #e5e7eb !important;
    font-family: 'Inter', sans-serif !important;
    font-size: 13.5px !important;
}
.stTextInput > div > div > input:focus,
.stTextArea > div > div > textarea:focus {
    border-color: #22c55e !important;
    box-shadow: 0 0 0 3px rgba(34,197,94,.1) !important;
}

/* Cards */
.ds-card {
    background: #fff;
    border-radius: 12px;
    border: 1px solid #e5e7eb;
    padding: 20px;
    margin-bottom: 16px;
    box-shadow: 0 1px 3px rgba(0,0,0,.05);
}
.ds-card-title {
    font-size: 14px;
    font-weight: 700;
    color: #111827;
    margin-bottom: 14px;
    padding-bottom: 10px;
    border-bottom: 1px solid #f3f4f6;
}

/* Found card */
.found-card {
    border: 1px solid #bbf7d0;
    border-radius: 10px;
    padding: 14px 16px;
    background: #f0fdf4;
    margin: 10px 0;
}
.found-card .reg {
    display: inline-block;
    font-size: 11px;
    font-weight: 700;
    color: #16a34a;
    background: #dcfce7;
    padding: 2px 8px;
    border-radius: 20px;
}
.found-card .fname {
    font-size: 15px;
    font-weight: 700;
    color: #111827;
    margin: 6px 0 2px;
}
.found-card .fmeta {
    font-size: 12.5px;
    color: #6b7280;
}

/* Badges */
.badge-issued {
    display: inline-block;
    padding: 2px 10px;
    border-radius: 20px;
    font-size: 11px;
    font-weight: 700;
    background: #fef3c7;
    color: #92400e;
    border: 1px solid #fcd34d;
}
.badge-returned {
    display: inline-block;
    padding: 2px 10px;
    border-radius: 20px;
    font-size: 11px;
    font-weight: 700;
    background: #dcfce7;
    color: #166534;
    border: 1px solid #86efac;
}

/* Section header */
.section-header {
    font-size: 18px;
    font-weight: 700;
    color: #111827;
    margin-bottom: 4px;
}
.section-sub {
    font-size: 13px;
    color: #6b7280;
    margin-bottom: 20px;
}

/* Alert styles */
.alert-success {
    padding: 10px 14px;
    border-radius: 8px;
    background: #f0fdf4;
    color: #16a34a;
    border: 1px solid #bbf7d0;
    font-size: 13px;
    font-weight: 500;
    margin: 8px 0;
}
.alert-warning {
    padding: 10px 14px;
    border-radius: 8px;
    background: #fffbeb;
    color: #d97706;
    border: 1px solid #fde68a;
    font-size: 13px;
    font-weight: 500;
    margin: 8px 0;
}
.alert-danger {
    padding: 10px 14px;
    border-radius: 8px;
    background: #fef2f2;
    color: #dc2626;
    border: 1px solid #fecaca;
    font-size: 13px;
    font-weight: 500;
    margin: 8px 0;
}

/* Dataframe */
[data-testid="stDataFrame"] {
    border-radius: 8px;
    border: 1px solid #e5e7eb;
}

/* Divider */
hr { border-color: #f3f4f6; }

/* Progress bar */
.stProgress > div > div > div > div {
    background: #16a34a;
}

/* Tab styling */
.stTabs [data-baseweb="tab-list"] {
    gap: 0;
    border-bottom: 1px solid #e5e7eb;
}
.stTabs [data-baseweb="tab"] {
    padding: 10px 18px;
    font-size: 13px;
    font-weight: 500;
    color: #6b7280;
    border-radius: 0;
}
.stTabs [aria-selected="true"] {
    color: #16a34a !important;
    border-bottom: 2px solid #16a34a !important;
    font-weight: 600 !important;
}

/* Upload zone */
[data-testid="stFileUploader"] {
    border: 2px dashed #d1fae5;
    border-radius: 12px;
    background: #f0fdf4;
    padding: 10px;
}

/* Search result list */
.sr-item {
    display: flex;
    align-items: center;
    gap: 10px;
    padding: 10px 14px;
    border-bottom: 1px solid #f3f4f6;
    border-radius: 8px;
    cursor: pointer;
    transition: background .12s;
}
.sr-item:hover { background: #f0fdf4; }
.sr-avatar {
    width: 36px; height: 36px;
    border-radius: 50%;
    background: linear-gradient(135deg,#22c55e,#16a34a);
    display: flex; align-items: center; justify-content: center;
    color: #fff; font-weight: 700; font-size: 13px;
    flex-shrink: 0;
}

/* Sidebar logo area */
.sidebar-logo {
    display: flex;
    align-items: center;
    gap: 10px;
    padding: 8px 0 20px;
    border-bottom: 1px solid #f3f4f6;
    margin-bottom: 16px;
}
.logo-box {
    width: 36px; height: 36px;
    background: linear-gradient(135deg,#22c55e,#16a34a);
    border-radius: 8px;
    display: flex; align-items: center; justify-content: center;
    color: #fff; font-weight: 700; font-size: 15px;
}
.logo-text {
    font-size: 15px;
    font-weight: 700;
    color: #111827;
}

/* Hide streamlit branding */
#MainMenu {visibility: hidden;}
footer {visibility: hidden;}
header {visibility: hidden;}
</style>
""", unsafe_allow_html=True)

# ── DATABASE ───────────────────────────────────────────────
DB_PATH = "distribution.db"

def get_db():
    conn = sqlite3.connect(DB_PATH, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = get_db(); c = conn.cursor()
    c.executescript("""
        CREATE TABLE IF NOT EXISTS students (
            sno        INTEGER PRIMARY KEY AUTOINCREMENT,
            reg_no     TEXT UNIQUE NOT NULL,
            name       TEXT NOT NULL,
            department TEXT NOT NULL,
            section    TEXT NOT NULL,
            created_at TEXT DEFAULT (datetime('now','localtime'))
        );
        CREATE TABLE IF NOT EXISTS transactions (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            reg_no      TEXT NOT NULL,
            name        TEXT NOT NULL,
            department  TEXT NOT NULL,
            section     TEXT NOT NULL,
            ipad_no     TEXT NOT NULL,
            stylus_no   TEXT NOT NULL,
            issued_at   TEXT DEFAULT (datetime('now','localtime')),
            returned_at TEXT,
            status      TEXT DEFAULT 'issued',
            condition   TEXT,
            remarks     TEXT,
            issued_by   TEXT DEFAULT 'Staff',
            FOREIGN KEY (reg_no) REFERENCES students(reg_no)
        );
    """)
    conn.commit(); conn.close()

init_db()

def db_query(sql, params=(), fetch="all"):
    conn = get_db()
    try:
        c = conn.cursor()
        c.execute(sql, params)
        if fetch == "all":
            return [dict(r) for r in c.fetchall()]
        elif fetch == "one":
            r = c.fetchone()
            return dict(r) if r else None
        elif fetch == "scalar":
            r = c.fetchone()
            return r[0] if r else 0
        elif fetch == "none":
            conn.commit()
            return None
    finally:
        conn.close()

def db_write(sql, params=()):
    conn = get_db()
    try:
        c = conn.cursor()
        c.execute(sql, params)
        conn.commit()
        return True
    except sqlite3.IntegrityError as e:
        return str(e)
    finally:
        conn.close()

# ── HELPERS ────────────────────────────────────────────────
def initials(name):
    return "".join(w[0] for w in name.split() if w).upper()[:2]

def card(title, body_fn):
    st.markdown(f'<div class="ds-card"><div class="ds-card-title">{title}</div></div>', unsafe_allow_html=True)

def found_card(s, extra=""):
    st.markdown(f"""
    <div class="found-card">
        <span class="reg">{s['reg_no']}</span>
        <div class="fname">{s['name']}</div>
        <div class="fmeta">🏛️ {s['department']} &nbsp;·&nbsp; 📌 Section {s['section']}</div>
        {extra}
    </div>""", unsafe_allow_html=True)

def badge(status):
    if status == "issued":
        return '<span class="badge-issued">Issued</span>'
    return '<span class="badge-returned">Returned</span>'

def style_header(ws, headers, color="16A34A"):
    fill   = PatternFill("solid", fgColor=color)
    font   = Font(bold=True, color="FFFFFF", size=11)
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'),  bottom=Side(style='thin'))
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = fill; cell.font = font
        cell.alignment = Alignment(horizontal='center'); cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = max(len(h)+4, 14)

# ── SIDEBAR ────────────────────────────────────────────────
with st.sidebar:
    st.markdown("""
    <div class="sidebar-logo">
        <div class="logo-box">DS</div>
        <div class="logo-text">Distribution System</div>
    </div>
    """, unsafe_allow_html=True)

    st.markdown("**MAIN**")
    page = st.radio(
        label="Navigation",
        options=["📤 Issue Devices", "📥 Return Devices", "📊 Dashboard",
                 "📋 Records", "🎓 Students", "📄 Reports"],
        label_visibility="collapsed"
    )
    st.markdown("---")
    st.markdown("""
    <div style="display:flex;align-items:center;gap:8px;padding:8px 0">
        <div style="width:32px;height:32px;border-radius:50%;background:linear-gradient(135deg,#22c55e,#16a34a);
             display:flex;align-items:center;justify-content:center;color:#fff;font-weight:700;font-size:12px">AD</div>
        <div>
            <div style="font-size:13px;font-weight:600;color:#111827">Admin</div>
            <div style="font-size:11px;color:#9ca3af">Staff</div>
        </div>
    </div>
    """, unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════
# PAGE: ISSUE DEVICES
# ══════════════════════════════════════════════════════════
if page == "📤 Issue Devices":
    st.markdown('<div class="section-header">📤 Issue Devices</div>', unsafe_allow_html=True)
    st.markdown('<div class="section-sub">Search a student and assign iPad & Stylus for the exam</div>', unsafe_allow_html=True)

    col1, col2 = st.columns(2)

    with col1:
        st.markdown('<div class="ds-card-title">🔍 Search Student</div>', unsafe_allow_html=True)
        q = st.text_input("Register No. or Name", placeholder="Type reg no or name…", key="issue_q")

        selected_student = None
        if q:
            results = db_query(
                "SELECT * FROM students WHERE reg_no LIKE ? OR name LIKE ? ORDER BY name LIMIT 20",
                (f"%{q}%", f"%{q}%")
            )
            if not results:
                st.markdown('<div class="alert-danger">❌ No student found.</div>', unsafe_allow_html=True)
            else:
                st.markdown(f"**{len(results)} student(s) found — select one:**")
                for s in results:
                    active = db_query(
                        "SELECT * FROM transactions WHERE reg_no=? AND status='issued'",
                        (s["reg_no"],), fetch="one"
                    )
                    status_label = "🟡 Issued" if active else "🟢 Available"
                    label = f"{initials(s['name'])}  |  {s['name']}  ·  {s['reg_no']}  ·  {s['department']} Sec {s['section']}  {status_label}"
                    if st.button(label, key=f"issue_sel_{s['reg_no']}"):
                        st.session_state["issue_student"] = s
                        st.session_state["issue_active"] = active
                        st.rerun()

        # Show selected student
        if "issue_student" in st.session_state and st.session_state["issue_student"]:
            s = st.session_state["issue_student"]
            active = st.session_state.get("issue_active")
            found_card(s)
            if active:
                st.markdown(f'<div class="alert-warning">⚠️ Already has active issue — iPad: {active["ipad_no"]}, Stylus: {active["stylus_no"]}</div>', unsafe_allow_html=True)
            if st.button("✕ Clear Selection", key="issue_clear"):
                st.session_state.pop("issue_student", None)
                st.session_state.pop("issue_active", None)
                st.rerun()

    with col2:
        st.markdown('<div class="ds-card-title">📲 Assign Devices</div>', unsafe_allow_html=True)
        s = st.session_state.get("issue_student")
        active = st.session_state.get("issue_active")
        disabled = not s or bool(active)

        ipad_no   = st.text_input("iPad Number",   placeholder="e.g. IPAD-042", disabled=disabled, key="ipad_no")
        stylus_no = st.text_input("Stylus Number", placeholder="e.g. STY-018",  disabled=disabled, key="stylus_no")
        issued_by = st.text_input("Issued By",     value="Staff",                disabled=disabled, key="issued_by")

        if st.button("✅ Issue Devices", disabled=disabled, key="issue_btn"):
            if not ipad_no or not stylus_no:
                st.error("Enter both iPad and Stylus numbers.")
            else:
                result = db_write(
                    "INSERT INTO transactions (reg_no,name,department,section,ipad_no,stylus_no,issued_by) VALUES (?,?,?,?,?,?,?)",
                    (s["reg_no"], s["name"], s["department"], s["section"], ipad_no, stylus_no, issued_by or "Staff")
                )
                if result is True:
                    st.success(f"✅ Issued to {s['name']} — iPad: {ipad_no}, Stylus: {stylus_no}")
                    st.session_state.pop("issue_student", None)
                    st.session_state.pop("issue_active", None)
                    st.rerun()
                else:
                    st.error(f"❌ Error: {result}")

# ══════════════════════════════════════════════════════════
# PAGE: RETURN DEVICES
# ══════════════════════════════════════════════════════════
elif page == "📥 Return Devices":
    st.markdown('<div class="section-header">📥 Return Devices</div>', unsafe_allow_html=True)
    st.markdown('<div class="section-sub">Find the active issue and mark devices as returned</div>', unsafe_allow_html=True)

    col1, col2 = st.columns(2)

    with col1:
        st.markdown('<div class="ds-card-title">🔍 Find Active Issue</div>', unsafe_allow_html=True)
        q = st.text_input("Register No. or Name", placeholder="Type reg no or name…", key="return_q")

        if q:
            results = db_query(
                "SELECT * FROM transactions WHERE status='issued' AND (reg_no LIKE ? OR name LIKE ?) ORDER BY issued_at DESC LIMIT 20",
                (f"%{q}%", f"%{q}%")
            )
            if not results:
                st.markdown('<div class="alert-danger">❌ No active issue found for this search.</div>', unsafe_allow_html=True)
            else:
                st.markdown(f"**{len(results)} active issue(s) — select one:**")
                for r in results:
                    label = f"{initials(r['name'])}  |  {r['name']}  ·  {r['reg_no']}  ·  iPad: {r['ipad_no']}  ·  Stylus: {r['stylus_no']}"
                    if st.button(label, key=f"ret_sel_{r['id']}"):
                        st.session_state["return_record"] = r
                        st.rerun()

        if "return_record" in st.session_state and st.session_state["return_record"]:
            r = st.session_state["return_record"]
            st.markdown(f"""
            <div class="found-card">
                <span class="reg">{r['reg_no']}</span>
                <span class="badge-issued" style="margin-left:8px">ISSUED</span>
                <div class="fname">{r['name']}</div>
                <div class="fmeta">🏛️ {r['department']} &nbsp;·&nbsp; 📌 Sec {r['section']}</div>
                <div style="margin-top:8px;font-size:12.5px;color:#374151">
                    📱 iPad: <b>{r['ipad_no']}</b> &nbsp;&nbsp; ✏️ Stylus: <b>{r['stylus_no']}</b>
                </div>
                <div style="font-size:11.5px;color:#9ca3af;margin-top:4px">Issued: {r['issued_at']}</div>
            </div>""", unsafe_allow_html=True)
            if st.button("✕ Clear Selection", key="return_clear"):
                st.session_state.pop("return_record", None)
                st.rerun()

    with col2:
        st.markdown('<div class="ds-card-title">📦 Process Return</div>', unsafe_allow_html=True)
        r = st.session_state.get("return_record")
        disabled = not r

        condition = st.selectbox("Device Condition",
            ["Good", "Minor Damage", "Major Damage", "Missing Stylus", "Missing iPad"],
            disabled=disabled, key="ret_condition")
        remarks = st.text_area("Remarks (optional)", placeholder="Any notes…",
            disabled=disabled, key="ret_remarks")

        st.markdown('<div class="warn-btn">', unsafe_allow_html=True)
        if st.button("↩️ Mark as Returned", disabled=disabled, key="return_btn"):
            db_write(
                "UPDATE transactions SET status='returned',returned_at=datetime('now','localtime'),condition=?,remarks=? WHERE id=? AND status='issued'",
                (condition, remarks, r["id"])
            )
            st.success(f"✅ Return processed for {r['name']} — iPad: {r['ipad_no']}")
            st.session_state.pop("return_record", None)
            st.rerun()
        st.markdown('</div>', unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════
# PAGE: DASHBOARD
# ══════════════════════════════════════════════════════════
elif page == "📊 Dashboard":
    st.markdown('<div class="section-header">📊 Dashboard</div>', unsafe_allow_html=True)
    st.markdown('<div class="section-sub">Live overview of device distribution</div>', unsafe_allow_html=True)

    total     = db_query("SELECT COUNT(*) FROM transactions",                           fetch="scalar")
    issued    = db_query("SELECT COUNT(*) FROM transactions WHERE status='issued'",     fetch="scalar")
    returned  = db_query("SELECT COUNT(*) FROM transactions WHERE status='returned'",   fetch="scalar")
    students  = db_query("SELECT COUNT(*) FROM students",                               fetch="scalar")

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("📋 Total Transactions", total)
    c2.metric("📤 Currently Issued",   issued)
    c3.metric("📥 Returned",           returned)
    c4.metric("🎓 Total Students",     students)

    st.markdown("---")
    col1, col2 = st.columns(2)

    with col1:
        st.markdown("**📊 Department Breakdown**")
        by_dept = db_query("""
            SELECT department,
                SUM(CASE WHEN status='issued'   THEN 1 ELSE 0 END) issued,
                SUM(CASE WHEN status='returned' THEN 1 ELSE 0 END) returned,
                COUNT(*) total
            FROM transactions GROUP BY department ORDER BY department
        """)
        if by_dept:
            df_dept = pd.DataFrame(by_dept)
            df_dept["Return %"] = (df_dept["returned"] / df_dept["total"] * 100).round(1)
            df_dept.columns = ["Department", "Issued", "Returned", "Total", "Return %"]
            st.dataframe(df_dept, use_container_width=True, hide_index=True)
        else:
            st.info("No transaction data yet.")

    with col2:
        st.markdown("**⏰ Recent Activity**")
        recent = db_query("SELECT * FROM transactions ORDER BY id DESC LIMIT 8")
        if recent:
            for r in recent:
                b = "🟡 Issued" if r["status"] == "issued" else "🟢 Returned"
                st.markdown(f"""
                <div style="display:flex;align-items:center;gap:10px;padding:9px 0;border-bottom:1px solid #f3f4f6">
                    <div style="width:34px;height:34px;border-radius:50%;background:linear-gradient(135deg,#22c55e,#16a34a);
                         display:flex;align-items:center;justify-content:center;color:#fff;font-weight:700;font-size:12px;flex-shrink:0">
                         {initials(r['name'])}</div>
                    <div style="flex:1">
                        <div style="font-size:13px;font-weight:600;color:#111827">{r['name']}</div>
                        <div style="font-size:12px;color:#9ca3af">{r['reg_no']} · iPad: {r['ipad_no']}</div>
                    </div>
                    <div style="font-size:11px;font-weight:600">{b}</div>
                </div>""", unsafe_allow_html=True)
        else:
            st.info("No activity yet.")

# ══════════════════════════════════════════════════════════
# PAGE: RECORDS
# ══════════════════════════════════════════════════════════
elif page == "📋 Records":
    st.markdown('<div class="section-header">📋 Transaction Records</div>', unsafe_allow_html=True)
    st.markdown('<div class="section-sub">All device issue and return history</div>', unsafe_allow_html=True)

    depts = [r["department"] for r in db_query("SELECT DISTINCT department FROM students ORDER BY department")]

    f1, f2, f3 = st.columns([1, 1, 2])
    with f1:
        f_status = st.selectbox("Status", ["All", "issued", "returned"], key="rec_status")
    with f2:
        f_dept = st.selectbox("Department", ["All"] + depts, key="rec_dept")
    with f3:
        f_q = st.text_input("Search name / reg no", placeholder="Search…", key="rec_q")

    sql = "SELECT * FROM transactions WHERE 1=1"; params = []
    if f_status != "All":  sql += " AND status=?";                          params.append(f_status)
    if f_dept   != "All":  sql += " AND department=?";                      params.append(f_dept)
    if f_q:                sql += " AND (reg_no LIKE ? OR name LIKE ?)";    params += [f"%{f_q}%", f"%{f_q}%"]
    sql += " ORDER BY id DESC"

    rows = db_query(sql, params)
    st.caption(f"{len(rows)} records")

    if rows:
        df = pd.DataFrame(rows)
        df = df[["reg_no","name","department","section","ipad_no","stylus_no",
                  "issued_at","returned_at","status","condition","remarks","issued_by"]]
        df.columns = ["Reg No","Name","Dept","Sec","iPad","Stylus",
                      "Issued At","Returned At","Status","Condition","Remarks","Issued By"]
        df["Returned At"] = df["Returned At"].fillna("—")
        df["Condition"]   = df["Condition"].fillna("—")
        df["Remarks"]     = df["Remarks"].fillna("—")
        st.dataframe(df, use_container_width=True, hide_index=True)
    else:
        st.info("No records found.")

# ══════════════════════════════════════════════════════════
# PAGE: STUDENTS
# ══════════════════════════════════════════════════════════
elif page == "🎓 Students":
    st.markdown('<div class="section-header">🎓 Students</div>', unsafe_allow_html=True)
    st.markdown('<div class="section-sub">Add individual students or bulk import from CSV, then browse the directory</div>', unsafe_allow_html=True)

    col1, col2 = st.columns(2)

    with col1:
        tab_add, tab_bulk = st.tabs(["➕ Add Student", "📁 Bulk Upload"])

        # ── ADD TAB ──
        with tab_add:
            with st.form("add_student_form", clear_on_submit=True):
                reg_no = st.text_input("Register Number *", placeholder="e.g. CS2024001")
                name   = st.text_input("Full Name *",       placeholder="Student full name")
                dept   = st.text_input("Department *",      placeholder="e.g. Computer Science")
                sec    = st.text_input("Section *",         placeholder="e.g. A")
                submitted = st.form_submit_button("Add Student", use_container_width=True)

            if submitted:
                if not all([reg_no, name, dept, sec]):
                    st.error("All fields are required.")
                else:
                    result = db_write(
                        "INSERT INTO students (reg_no,name,department,section) VALUES (?,?,?,?)",
                        (reg_no.strip(), name.strip(), dept.strip(), sec.strip())
                    )
                    if result is True:
                        st.success(f"✅ Student {name} added!")
                        st.rerun()
                    else:
                        st.error(f"❌ Register number already exists.")

        # ── BULK UPLOAD TAB ──
        with tab_bulk:
            st.markdown("**CSV Format:**")
            st.code("sno,reg_no,name,department,section\n1,CS2024001,Arjun Mehta,Computer Science,A\n2,EC2024001,Priya Sharma,Electronics,B")

            # Template download
            template_csv = "sno,reg_no,name,department,section\n1,CS2024001,Student Name,Computer Science,A\n2,EC2024001,Student Name,Electronics,B"
            st.download_button("⬇️ Download Template", template_csv, "student_template.csv", "text/csv")

            uploaded = st.file_uploader("Upload CSV File", type=["csv"], key="bulk_csv")
            if uploaded:
                try:
                    df_upload = pd.read_csv(uploaded)
                    df_upload.columns = [c.strip().lower() for c in df_upload.columns]
                    required = ["sno","reg_no","name","department","section"]
                    missing  = [c for c in required if c not in df_upload.columns]
                    if missing:
                        st.error(f"❌ Missing columns: {', '.join(missing)}")
                    else:
                        st.markdown(f"**Preview — {len(df_upload)} rows:**")
                        st.dataframe(df_upload.head(10), use_container_width=True, hide_index=True)

                        if st.button("✅ Confirm & Upload", key="confirm_upload"):
                            inserted = skipped = 0
                            progress = st.progress(0)
                            total_rows = len(df_upload)
                            for i, row in df_upload.iterrows():
                                result = db_write(
                                    "INSERT INTO students (reg_no,name,department,section) VALUES (?,?,?,?)",
                                    (str(row["reg_no"]).strip(), str(row["name"]).strip(),
                                     str(row["department"]).strip(), str(row["section"]).strip())
                                )
                                if result is True: inserted += 1
                                else:              skipped  += 1
                                progress.progress((i+1)/total_rows)
                            st.success(f"✅ Upload complete — {inserted} added, {skipped} skipped (duplicates).")
                            st.rerun()
                except Exception as e:
                    st.error(f"❌ Error reading CSV: {e}")

    with col2:
        st.markdown("**All Students**")
        search_s = st.text_input("Search students", placeholder="Name or reg no…", key="stu_search")
        students_list = db_query(
            "SELECT * FROM students WHERE name LIKE ? OR reg_no LIKE ? ORDER BY sno",
            (f"%{search_s}%", f"%{search_s}%")
        )
        st.caption(f"{len(students_list)} students")
        for s in students_list:
            st.markdown(f"""
            <div style="display:flex;align-items:center;gap:10px;padding:10px 0;border-bottom:1px solid #f3f4f6">
                <div style="width:36px;height:36px;border-radius:50%;background:linear-gradient(135deg,#22c55e,#16a34a);
                     display:flex;align-items:center;justify-content:center;color:#fff;font-weight:700;font-size:13px;flex-shrink:0">
                     {initials(s['name'])}</div>
                <div style="flex:1">
                    <div style="font-size:13.5px;font-weight:600;color:#111827">{s['name']}</div>
                    <div style="font-size:12px;color:#9ca3af">{s['reg_no']} · {s['department']} · Sec {s['section']}</div>
                </div>
                <div style="font-size:11px;color:#9ca3af">#{s['sno']}</div>
            </div>""", unsafe_allow_html=True)

        if not students_list:
            st.info("No students found.")

# ══════════════════════════════════════════════════════════
# PAGE: REPORTS
# ══════════════════════════════════════════════════════════
elif page == "📄 Reports":
    st.markdown('<div class="section-header">📄 Reports & Export</div>', unsafe_allow_html=True)
    st.markdown('<div class="section-sub">Download data in Excel or CSV format</div>', unsafe_allow_html=True)

    # DB Info
    total_s = db_query("SELECT COUNT(*) FROM students",                         fetch="scalar")
    total_t = db_query("SELECT COUNT(*) FROM transactions",                     fetch="scalar")
    issued  = db_query("SELECT COUNT(*) FROM transactions WHERE status='issued'",fetch="scalar")
    size_kb = os.path.getsize(DB_PATH)//1024 if os.path.exists(DB_PATH) else 0

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("🗄️ Database",      "SQLite")
    c2.metric("💾 File Size",     f"{size_kb} KB")
    c3.metric("🎓 Students",      total_s)
    c4.metric("📋 Transactions",  total_t)

    st.markdown("---")

    # Filters
    st.markdown("**⚙️ Filter Options**")
    depts = [r["department"] for r in db_query("SELECT DISTINCT department FROM students ORDER BY department")]
    f1, f2, f3, f4 = st.columns(4)
    with f1: r_from   = st.date_input("From Date", value=None, key="r_from")
    with f2: r_to     = st.date_input("To Date",   value=None, key="r_to")
    with f3: r_dept   = st.selectbox("Department", ["All"] + depts, key="r_dept")
    with f4: r_status = st.selectbox("Status", ["All","issued","returned"], key="r_status")

    def get_filtered():
        sql = "SELECT * FROM transactions WHERE 1=1"; params = []
        if r_from:               sql += " AND DATE(issued_at)>=?"; params.append(str(r_from))
        if r_to:                 sql += " AND DATE(issued_at)<=?"; params.append(str(r_to))
        if r_dept   != "All":   sql += " AND department=?";        params.append(r_dept)
        if r_status != "All":   sql += " AND status=?";            params.append(r_status)
        return db_query(sql + " ORDER BY id DESC", params)

    st.markdown("---")
    st.markdown("**📥 Download Reports**")

    r1, r2, r3 = st.columns(3)
    r4, r5, r6 = st.columns(3)

    # ── Transactions Excel ──
    with r1:
        st.markdown("**📊 Full Transaction Report**")
        st.caption("Excel (.xlsx) — All records")
        if st.button("⬇️ Download", key="dl_txn_xl"):
            rows = get_filtered()
            wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Transactions"
            hdrs = ["S.No","Reg No","Name","Department","Section","iPad No","Stylus No",
                    "Issued At","Returned At","Status","Condition","Remarks","Issued By"]
            style_header(ws, hdrs)
            alt = PatternFill("solid", fgColor="F0FDF4")
            for i, r in enumerate(rows, 1):
                row_data = [i,r["reg_no"],r["name"],r["department"],r["section"],r["ipad_no"],
                            r["stylus_no"],r["issued_at"],r.get("returned_at") or "",r["status"],
                            r.get("condition") or "",r.get("remarks") or "",r.get("issued_by") or ""]
                for col, val in enumerate(row_data, 1):
                    cell = ws.cell(row=i+1, column=col, value=val)
                    if i%2==0: cell.fill = alt
            buf = io.BytesIO(); wb.save(buf); buf.seek(0)
            st.download_button("💾 Save Excel", buf, "transactions_report.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="save_txn_xl")

    # ── Students Excel ──
    with r2:
        st.markdown("**🎓 Student Directory**")
        st.caption("Excel (.xlsx) — All students")
        if st.button("⬇️ Download", key="dl_stu_xl"):
            rows = db_query("SELECT * FROM students ORDER BY sno")
            wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Students"
            style_header(ws, ["S.No","Reg No","Name","Department","Section","Added At"], "0F172A")
            alt = PatternFill("solid", fgColor="F9FAFB")
            for i, r in enumerate(rows, 1):
                row_data = [r["sno"],r["reg_no"],r["name"],r["department"],r["section"],r["created_at"]]
                for col, val in enumerate(row_data, 1):
                    cell = ws.cell(row=i+1, column=col, value=val)
                    if i%2==0: cell.fill = alt
            buf = io.BytesIO(); wb.save(buf); buf.seek(0)
            st.download_button("💾 Save Excel", buf, "students_directory.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="save_stu_xl")

    # ── Summary Excel ──
    with r3:
        st.markdown("**📈 Summary Report**")
        st.caption("Excel (.xlsx) — Dept statistics")
        if st.button("⬇️ Download", key="dl_sum_xl"):
            by_dept = db_query("""SELECT department,COUNT(*) total_transactions,
                SUM(CASE WHEN status='issued' THEN 1 ELSE 0 END) currently_issued,
                SUM(CASE WHEN status='returned' THEN 1 ELSE 0 END) returned,
                COUNT(DISTINCT reg_no) unique_students FROM transactions GROUP BY department""")
            wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Dept Summary"
            style_header(ws, ["Department","Total Transactions","Currently Issued","Returned","Unique Students"])
            alt = PatternFill("solid", fgColor="F0FDF4")
            for i, r in enumerate(by_dept, 1):
                row_data = [r["department"],r["total_transactions"],r["currently_issued"],r["returned"],r["unique_students"]]
                for col, val in enumerate(row_data, 1):
                    cell = ws.cell(row=i+1, column=col, value=val)
                    if i%2==0: cell.fill = alt
            buf = io.BytesIO(); wb.save(buf); buf.seek(0)
            st.download_button("💾 Save Excel", buf, "summary_report.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="save_sum_xl")

    # ── Transactions CSV ──
    with r4:
        st.markdown("**📋 Transactions CSV**")
        st.caption("Raw CSV — All transactions")
        if st.button("⬇️ Download", key="dl_txn_csv"):
            rows = get_filtered()
            out = io.StringIO(); w = csv.writer(out)
            w.writerow(["S.No","Reg No","Name","Department","Section","iPad No","Stylus No",
                        "Issued At","Returned At","Status","Condition","Remarks","Issued By"])
            for i, r in enumerate(rows, 1):
                w.writerow([i,r["reg_no"],r["name"],r["department"],r["section"],r["ipad_no"],r["stylus_no"],
                            r["issued_at"],r.get("returned_at") or "",r["status"],
                            r.get("condition") or "",r.get("remarks") or "",r.get("issued_by") or ""])
            st.download_button("💾 Save CSV", out.getvalue(), "transactions.csv", "text/csv", key="save_txn_csv")

    # ── Students CSV ──
    with r5:
        st.markdown("**📄 Students CSV**")
        st.caption("Raw CSV — Student list")
        if st.button("⬇️ Download", key="dl_stu_csv"):
            rows = db_query("SELECT * FROM students ORDER BY sno")
            out = io.StringIO(); w = csv.writer(out)
            w.writerow(["S.No","Reg No","Name","Department","Section","Added At"])
            for r in rows:
                w.writerow([r["sno"],r["reg_no"],r["name"],r["department"],r["section"],r["created_at"]])
            st.download_button("💾 Save CSV", out.getvalue(), "students.csv", "text/csv", key="save_stu_csv")

    # ── Currently Issued CSV ──
    with r6:
        st.markdown("**⚠️ Currently Issued**")
        st.caption("CSV — Pending returns")
        if st.button("⬇️ Download", key="dl_iss_csv"):
            rows = db_query("SELECT * FROM transactions WHERE status='issued' ORDER BY issued_at")
            out = io.StringIO(); w = csv.writer(out)
            w.writerow(["S.No","Reg No","Name","Department","Section","iPad No","Stylus No","Issued At","Issued By"])
            for i, r in enumerate(rows, 1):
                w.writerow([i,r["reg_no"],r["name"],r["department"],r["section"],
                            r["ipad_no"],r["stylus_no"],r["issued_at"],r.get("issued_by") or ""])
            st.download_button("💾 Save CSV", out.getvalue(), "currently_issued.csv", "text/csv", key="save_iss_csv")
