"""
iPad & Stylus Distribution Management System
Flask + PostgreSQL — heycampus-style UI
Changes:
  • Students page: Add + Bulk Upload combined with inner tabs
  • Issue / Return search: lists ALL matching profiles; click to select
"""

from flask import Flask, request, jsonify, render_template_string, send_file
import psycopg2
import psycopg2.extras
from psycopg2 import errors as pg_errors
import csv, io, os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)

# ── DB CONFIG — edit these or set environment variables ──
DB_CONFIG = {
    "host":     os.getenv("PG_HOST",     "localhost"),
    "port":     int(os.getenv("PG_PORT", "5432")),
    "dbname":   os.getenv("PG_DB",       "distribution_db"),
    "user":     os.getenv("PG_USER",     "postgres"),
    "password": os.getenv("PG_PASSWORD", "yourpassword"),
}

# ── DB HELPERS ────────────────────────────────────────────

def get_db():
    """Open a new PostgreSQL connection."""
    return psycopg2.connect(**DB_CONFIG)

def query(sql_str, params=None, fetch="all"):
    """
    Run SQL and return results.
      fetch='all'  → list of dicts
      fetch='one'  → single dict or None
      fetch='none' → nothing (INSERT / UPDATE / DELETE)
    """
    conn = get_db()
    try:
        with conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                cur.execute(sql_str, params)
                if fetch == "all":
                    return [dict(r) for r in cur.fetchall()]
                elif fetch == "one":
                    row = cur.fetchone()
                    return dict(row) if row else None
                elif fetch == "none":
                    return None
    finally:
        conn.close()

def scalar(sql_str, params=None):
    """Return a single scalar value."""
    conn = get_db()
    try:
        with conn:
            with conn.cursor() as cur:
                cur.execute(sql_str, params)
                row = cur.fetchone()
                return row[0] if row else 0
    finally:
        conn.close()

def ts(d, *keys):
    """Stringify timestamp fields in a dict so JSON serialisation works."""
    if d:
        for k in keys:
            if d.get(k):
                d[k] = str(d[k])
    return d

# ── DATABASE SETUP ────────────────────────────────────────

def init_db():
    conn = get_db()
    try:
        with conn:
            with conn.cursor() as c:
                c.execute("""
                    CREATE TABLE IF NOT EXISTS students (
                        sno        SERIAL PRIMARY KEY,
                        reg_no     TEXT UNIQUE NOT NULL,
                        name       TEXT NOT NULL,
                        department TEXT NOT NULL,
                        section    TEXT NOT NULL,
                        created_at TIMESTAMP DEFAULT NOW()
                    );
                """)
                c.execute("""
                    CREATE TABLE IF NOT EXISTS transactions (
                        id          SERIAL PRIMARY KEY,
                        reg_no      TEXT NOT NULL REFERENCES students(reg_no),
                        name        TEXT NOT NULL,
                        department  TEXT NOT NULL,
                        section     TEXT NOT NULL,
                        ipad_no     TEXT NOT NULL,
                        stylus_no   TEXT NOT NULL,
                        issued_at   TIMESTAMP DEFAULT NOW(),
                        returned_at TIMESTAMP,
                        status      TEXT DEFAULT 'issued',
                        condition   TEXT,
                        remarks     TEXT,
                        issued_by   TEXT DEFAULT 'Staff'
                    );
                """)
        print("✅ PostgreSQL tables ready.")
    except Exception as e:
        print(f"⚠️  DB init warning: {e}")
    finally:
        conn.close()

init_db()

# ── HTML (same heycampus-style UI) ───────────────────────


HTML = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>iPad & Stylus Distribution</title>
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap" rel="stylesheet">
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:'Inter',sans-serif;background:#f4f5f7;color:#111827;height:100vh;display:flex;flex-direction:column}

/* TOP BAR */
.topbar{height:56px;background:#fff;border-bottom:1px solid #e5e7eb;display:flex;align-items:center;justify-content:space-between;padding:0 20px;flex-shrink:0;position:sticky;top:0;z-index:100}
.topbar-left{display:flex;align-items:center;gap:12px}
.topbar-logo{width:32px;height:32px;background:linear-gradient(135deg,#22c55e,#16a34a);border-radius:8px;display:flex;align-items:center;justify-content:center;color:#fff;font-weight:700;font-size:14px}
.topbar-title{font-weight:700;font-size:15px;color:#111827}
.topbar-right{display:flex;align-items:center;gap:10px}
.topbar-icon{width:34px;height:34px;border-radius:50%;background:#f9fafb;border:1px solid #e5e7eb;display:flex;align-items:center;justify-content:center;cursor:pointer;font-size:15px}
.topbar-avatar{width:34px;height:34px;border-radius:50%;background:linear-gradient(135deg,#22c55e,#16a34a);display:flex;align-items:center;justify-content:center;color:#fff;font-weight:700;font-size:13px;cursor:pointer}

/* LAYOUT */
.app-body{display:flex;flex:1;overflow:hidden}

/* SIDEBAR */
.sidebar{width:220px;background:#fff;border-right:1px solid #e5e7eb;display:flex;flex-direction:column;flex-shrink:0;overflow-y:auto}
.sidebar-section-label{font-size:11px;font-weight:600;color:#9ca3af;text-transform:uppercase;letter-spacing:.7px;padding:20px 16px 8px}
.nav-item{display:flex;align-items:center;gap:10px;padding:9px 12px;margin:1px 8px;border-radius:8px;cursor:pointer;font-size:13.5px;font-weight:500;color:#374151;transition:all .15s;text-decoration:none}
.nav-item:hover{background:#f0fdf4;color:#16a34a}
.nav-item.active{background:#16a34a;color:#fff}
.nav-icon{font-size:16px;width:20px;text-align:center;color:#6b7280}
.nav-item.active .nav-icon{color:#fff}
.sidebar-bottom{margin-top:auto;border-top:1px solid #e5e7eb;padding:12px}
.sidebar-user{display:flex;align-items:center;gap:10px;padding:8px;border-radius:8px;cursor:pointer}
.sidebar-user:hover{background:#f9fafb}
.sidebar-user-avatar{width:32px;height:32px;border-radius:50%;background:linear-gradient(135deg,#22c55e,#16a34a);display:flex;align-items:center;justify-content:center;color:#fff;font-weight:700;font-size:12px;flex-shrink:0}
.sidebar-user-info .name{font-size:13px;font-weight:600;color:#111827}
.sidebar-user-info .role{font-size:11px;color:#9ca3af}

/* MAIN */
.main{flex:1;overflow-y:auto;padding:24px}
.page{display:none}
.page.active{display:block}
.page-header{margin-bottom:20px}
.page-header h1{font-size:18px;font-weight:700;color:#111827}
.page-header p{font-size:13px;color:#6b7280;margin-top:2px}

/* CARDS */
.card{background:#fff;border-radius:12px;border:1px solid #e5e7eb;box-shadow:0 1px 3px rgba(0,0,0,.05)}
.card-pad{padding:20px}
.card-title{font-size:14px;font-weight:700;color:#111827;margin-bottom:16px;padding-bottom:12px;border-bottom:1px solid #f3f4f6}

/* INNER TABS (inside a card) */
.inner-tabs{display:flex;border-bottom:1px solid #e5e7eb;margin-bottom:20px}
.inner-tab{padding:10px 18px;font-size:13px;font-weight:500;color:#6b7280;cursor:pointer;border-bottom:2px solid transparent;margin-bottom:-1px;transition:all .15s}
.inner-tab:hover{color:#16a34a}
.inner-tab.active{color:#16a34a;border-bottom-color:#16a34a;font-weight:600}
.inner-tab-pane{display:none}
.inner-tab-pane.active{display:block}

/* STAT GRID */
.stat-grid{display:grid;grid-template-columns:repeat(4,1fr);gap:16px;margin-bottom:20px}
.stat-card{background:#fff;border-radius:12px;border:1px solid #e5e7eb;padding:18px 20px;display:flex;align-items:center;gap:14px}
.stat-icon{width:44px;height:44px;border-radius:10px;display:flex;align-items:center;justify-content:center;font-size:20px;flex-shrink:0}
.stat-icon.green{background:#f0fdf4}.stat-icon.amber{background:#fffbeb}
.stat-icon.blue{background:#eff6ff}.stat-icon.purple{background:#faf5ff}
.stat-label{font-size:12px;color:#6b7280;font-weight:500}
.stat-value{font-size:22px;font-weight:700;color:#111827;margin-top:1px}

/* GRIDS */
.two-col{display:grid;grid-template-columns:1fr 1fr;gap:16px}
.three-col{display:grid;grid-template-columns:repeat(3,1fr);gap:16px}

/* PROFILE CARD */
.profile-card{background:#fff;border-radius:12px;border:1px solid #e5e7eb;padding:24px;text-align:center}
.profile-avatar{width:80px;height:80px;border-radius:16px;margin:0 auto 12px;display:flex;align-items:center;justify-content:center;font-size:28px;font-weight:700;color:#fff;background:linear-gradient(135deg,#22c55e,#16a34a)}
.profile-name{font-size:16px;font-weight:700;color:#111827}
.profile-sub{font-size:13px;color:#6b7280;margin-top:3px}

/* INFO PANEL */
.info-panel{background:#fff;border-radius:12px;border:1px solid #e5e7eb}
.info-tabs{display:flex;border-bottom:1px solid #e5e7eb;padding:0 20px}
.info-tab{padding:12px 16px;font-size:13.5px;font-weight:500;color:#6b7280;cursor:pointer;border-bottom:2px solid transparent;margin-bottom:-1px;transition:all .15s}
.info-tab.active{color:#16a34a;border-bottom-color:#16a34a;font-weight:600}
.info-tab-content{display:none;padding:20px}
.info-tab-content.active{display:block}
.info-section-title{font-size:14px;font-weight:700;color:#111827;margin-bottom:16px}
.info-grid{display:grid;grid-template-columns:repeat(3,1fr);gap:20px;margin-bottom:20px}
.info-field{display:flex;align-items:flex-start;gap:10px}
.info-field-icon{color:#9ca3af;font-size:16px;margin-top:1px;flex-shrink:0}
.info-field-label{font-size:11px;color:#9ca3af;font-weight:500;margin-bottom:2px}
.info-field-value{font-size:13.5px;color:#111827;font-weight:500}

/* FORMS */
.form-group{margin-bottom:14px}
.form-label{display:block;font-size:12px;font-weight:600;color:#374151;margin-bottom:5px;letter-spacing:.2px}
.form-control{width:100%;border:1px solid #e5e7eb;border-radius:8px;padding:8px 12px;font-size:13.5px;color:#111827;outline:none;transition:border .15s;font-family:'Inter',sans-serif;background:#fff}
.form-control:focus{border-color:#22c55e;box-shadow:0 0 0 3px rgba(34,197,94,.1)}
.form-control:disabled{background:#f9fafb;color:#9ca3af;cursor:not-allowed}
textarea.form-control{resize:vertical}

/* BUTTONS */
.btn{display:inline-flex;align-items:center;gap:6px;padding:8px 16px;border-radius:8px;font-size:13.5px;font-weight:600;border:none;cursor:pointer;transition:all .15s;font-family:'Inter',sans-serif}
.btn-primary{background:#16a34a;color:#fff}.btn-primary:hover{background:#15803d}
.btn-warning{background:#f59e0b;color:#fff}.btn-warning:hover{background:#d97706}
.btn-outline{background:#fff;color:#374151;border:1px solid #e5e7eb}.btn-outline:hover{background:#f9fafb}
.btn-block{width:100%;justify-content:center}
.btn:disabled{opacity:.4;cursor:not-allowed}
.btn-sm{padding:5px 12px;font-size:12px}
.search-row{display:flex;gap:8px}
.search-row .form-control{flex:1}

/* ALERTS */
.alert{padding:10px 14px;border-radius:8px;font-size:13px;font-weight:500;margin-top:10px}
.alert-success{background:#f0fdf4;color:#16a34a;border:1px solid #bbf7d0}
.alert-warning{background:#fffbeb;color:#d97706;border:1px solid #fde68a}
.alert-danger{background:#fef2f2;color:#dc2626;border:1px solid #fecaca}
.alert-info{background:#eff6ff;color:#2563eb;border:1px solid #bfdbfe}

/* SEARCH RESULTS LIST */
.search-results{margin-top:10px;border:1px solid #e5e7eb;border-radius:10px;overflow:hidden;display:none}
.search-result-item{display:flex;align-items:center;gap:12px;padding:11px 14px;cursor:pointer;border-bottom:1px solid #f3f4f6;transition:background .12s}
.search-result-item:last-child{border-bottom:none}
.search-result-item:hover{background:#f0fdf4}
.search-result-item.selected{background:#dcfce7;border-left:3px solid #16a34a}
.sr-avatar{width:36px;height:36px;border-radius:50%;display:flex;align-items:center;justify-content:center;font-size:13px;font-weight:700;color:#fff;flex-shrink:0}
.sr-name{font-size:13.5px;font-weight:600;color:#111827}
.sr-meta{font-size:12px;color:#9ca3af;margin-top:1px}
.sr-badge{margin-left:auto;flex-shrink:0;font-size:11px;font-weight:700;padding:2px 8px;border-radius:20px}
.sr-badge.issued{background:#fef3c7;color:#92400e;border:1px solid #fcd34d}
.sr-badge.available{background:#dcfce7;color:#166534;border:1px solid #86efac}
.search-no-result{padding:16px;text-align:center;font-size:13px;color:#9ca3af}

/* FOUND CARD (selected student) */
.found-card{border:1px solid #bbf7d0;border-radius:10px;padding:14px;background:#f0fdf4;margin-top:10px;position:relative}
.found-reg{display:inline-block;font-size:11px;font-weight:700;color:#16a34a;background:#dcfce7;padding:2px 8px;border-radius:20px}
.found-name{font-size:15px;font-weight:700;color:#111827;margin:6px 0 4px}
.found-meta{font-size:12.5px;color:#6b7280;display:flex;gap:16px;flex-wrap:wrap}
.found-clear{position:absolute;top:10px;right:10px;background:none;border:none;color:#9ca3af;cursor:pointer;font-size:16px;line-height:1}
.found-clear:hover{color:#dc2626}

/* MEMBER LIST */
.member-item{display:flex;align-items:center;gap:10px;padding:10px 0;border-bottom:1px solid #f9fafb}
.member-item:last-child{border-bottom:none}
.member-avatar{width:36px;height:36px;border-radius:50%;display:flex;align-items:center;justify-content:center;font-size:13px;font-weight:700;color:#fff;flex-shrink:0}
.member-name{font-size:13.5px;font-weight:600;color:#111827}
.member-email{font-size:12px;color:#9ca3af}
.member-status{margin-left:auto;flex-shrink:0}

/* TABLE */
.table-wrap{overflow-x:auto}
table{width:100%;border-collapse:collapse;font-size:13px}
thead th{background:#f9fafb;color:#6b7280;font-weight:600;font-size:11.5px;text-transform:uppercase;letter-spacing:.4px;padding:10px 14px;text-align:left;border-bottom:1px solid #e5e7eb}
tbody td{padding:11px 14px;border-bottom:1px solid #f3f4f6;color:#374151}
tbody tr:hover td{background:#fafafa}
tbody tr:last-child td{border-bottom:none}
code{font-size:12px;background:#f3f4f6;padding:2px 6px;border-radius:4px;color:#374151}

/* FILTER BAR */
.filter-bar{display:flex;gap:10px;margin-bottom:16px;flex-wrap:wrap;align-items:center}
.filter-bar .form-control{width:auto;flex:1;min-width:130px}
.count-pill{font-size:12px;color:#6b7280;background:#f3f4f6;padding:4px 10px;border-radius:20px;white-space:nowrap}

/* UPLOAD */
.upload-zone{border:2px dashed #d1fae5;border-radius:12px;padding:36px;text-align:center;cursor:pointer;background:#f0fdf4;transition:all .15s}
.upload-zone:hover,.upload-zone.drag{border-color:#16a34a;background:#dcfce7}
.upload-zone-icon{font-size:36px;margin-bottom:10px}
.upload-zone p{font-size:13.5px;color:#6b7280}
.upload-zone strong{color:#16a34a}
#csvFile{display:none}
.csv-code{background:#0f172a;color:#94a3b8;padding:16px;border-radius:10px;font-family:monospace;font-size:12.5px;line-height:1.8;margin-top:12px}
.csv-code .ch{color:#7dd3fc}.csv-code .cr{color:#86efac}
.progress-wrap{display:none;margin-top:12px}
.progress-bar-bg{height:6px;background:#e5e7eb;border-radius:3px;overflow:hidden}
.progress-bar-fill{height:6px;background:#16a34a;border-radius:3px;transition:width .3s}

/* REPORTS */
.report-grid{display:grid;grid-template-columns:repeat(3,1fr);gap:12px}
.report-btn{display:flex;align-items:center;gap:12px;background:#fff;border:1px solid #e5e7eb;border-radius:10px;padding:14px 16px;cursor:pointer;transition:all .15s}
.report-btn:hover{border-color:#22c55e;background:#f0fdf4}
.report-btn-icon{font-size:24px}
.report-btn-title{font-size:13px;font-weight:700;color:#111827}
.report-btn-sub{font-size:11.5px;color:#9ca3af;margin-top:1px}
.section-divider{font-size:13px;font-weight:700;color:#111827;margin:20px 0 12px;display:flex;align-items:center;gap:8px}
.section-divider::after{content:'';flex:1;height:1px;background:#f3f4f6}

/* AVATAR COLORS */
.av-green{background:linear-gradient(135deg,#22c55e,#16a34a)}
.av-blue{background:linear-gradient(135deg,#3b82f6,#2563eb)}
.av-purple{background:linear-gradient(135deg,#a855f7,#7c3aed)}
.av-amber{background:linear-gradient(135deg,#f59e0b,#d97706)}
.av-rose{background:linear-gradient(135deg,#f43f5e,#e11d48)}
.av-cyan{background:linear-gradient(135deg,#06b6d4,#0891b2)}
.av-indigo{background:linear-gradient(135deg,#6366f1,#4f46e5)}
.av-teal{background:linear-gradient(135deg,#14b8a6,#0f766e)}

@media(max-width:900px){
  .sidebar{display:none}
  .stat-grid{grid-template-columns:1fr 1fr}
  .two-col,.info-grid{grid-template-columns:1fr}
  .three-col,.report-grid{grid-template-columns:1fr}
}
</style>
</head>
<body>

<!-- TOP BAR -->
<div class="topbar">
  <div class="topbar-left">
    <div class="topbar-logo">DS</div>
    <span class="topbar-title">Distribution System</span>
  </div>
  <div class="topbar-right">
    <div class="topbar-icon">☀️</div>
    <div class="topbar-icon">⚙️</div>
    <div class="topbar-icon">🔔</div>
    <div class="topbar-avatar">AD</div>
  </div>
</div>

<div class="app-body">
  <!-- SIDEBAR -->
  <div class="sidebar">
    <div class="sidebar-section-label">Main</div>
    <a class="nav-item active" onclick="nav('issue',this)"><span class="nav-icon">📤</span> Issue Devices</a>
    <a class="nav-item" onclick="nav('return',this)"><span class="nav-icon">📥</span> Return Devices</a>
    <a class="nav-item" onclick="nav('dashboard',this)"><span class="nav-icon">📊</span> Dashboard</a>
    <a class="nav-item" onclick="nav('records',this)"><span class="nav-icon">📋</span> Records</a>
    <a class="nav-item" onclick="nav('students',this)"><span class="nav-icon">🎓</span> Students</a>
    <a class="nav-item" onclick="nav('reports',this)"><span class="nav-icon">📄</span> Reports</a>
    <div class="sidebar-bottom">
      <div class="sidebar-user">
        <div class="sidebar-user-avatar">AD</div>
        <div class="sidebar-user-info">
          <div class="name">Admin</div>
          <div class="role">Staff</div>
        </div>
      </div>
    </div>
  </div>

  <div class="main">

    <!-- ══════════════════════════════════
         ISSUE PAGE
    ══════════════════════════════════ -->
    <div id="page-issue" class="page active">
      <div class="page-header">
        <h1>Issue Devices</h1>
        <p>Search a student and assign iPad & Stylus for the exam</p>
      </div>
      <div class="two-col">
        <div class="card card-pad">
          <div class="card-title">🔍 Search Student</div>
          <div class="form-group">
            <label class="form-label">Register No. or Name</label>
            <div class="search-row">
              <input id="issueSearch" class="form-control" placeholder="Type reg no or name…"
                oninput="searchStudentLive('issue')" onkeydown="if(event.key==='Enter') searchStudentLive('issue')">
              <button class="btn btn-primary" onclick="searchStudentLive('issue')">Search</button>
            </div>
          </div>
          <!-- Multi-result list -->
          <div class="search-results" id="issueResults"></div>
          <!-- Selected card -->
          <div id="issueStudentResult"></div>
        </div>

        <div class="card card-pad">
          <div class="card-title">📲 Assign Devices</div>
          <div class="form-group"><label class="form-label">iPad Number</label>
            <input id="ipadNo" class="form-control" placeholder="e.g. IPAD-042" disabled></div>
          <div class="form-group"><label class="form-label">Stylus Number</label>
            <input id="stylusNo" class="form-control" placeholder="e.g. STY-018" disabled></div>
          <div class="form-group"><label class="form-label">Issued By (Staff Name)</label>
            <input id="issuedBy" class="form-control" value="Staff" disabled></div>
          <button class="btn btn-primary btn-block" id="issueBtn" onclick="issueDevices()" disabled>
            ✅ Issue Devices
          </button>
          <div id="issueAlert"></div>
        </div>
      </div>
    </div>

    <!-- ══════════════════════════════════
         RETURN PAGE
    ══════════════════════════════════ -->
    <div id="page-return" class="page">
      <div class="page-header">
        <h1>Return Devices</h1>
        <p>Find the active issue and mark devices as returned</p>
      </div>
      <div class="two-col">
        <div class="card card-pad">
          <div class="card-title">🔍 Find Active Issue</div>
          <div class="form-group">
            <label class="form-label">Register No. or Name</label>
            <div class="search-row">
              <input id="returnSearch" class="form-control" placeholder="Type reg no or name…"
                oninput="searchReturnLive()" onkeydown="if(event.key==='Enter') searchReturnLive()">
              <button class="btn btn-primary" onclick="searchReturnLive()">Find</button>
            </div>
          </div>
          <div class="search-results" id="returnResults"></div>
          <div id="returnStudentResult"></div>
        </div>

        <div class="card card-pad">
          <div class="card-title">📦 Process Return</div>
          <div class="form-group">
            <label class="form-label">Device Condition</label>
            <select id="retCondition" class="form-control" disabled>
              <option>Good</option><option>Minor Damage</option>
              <option>Major Damage</option><option>Missing Stylus</option><option>Missing iPad</option>
            </select>
          </div>
          <div class="form-group"><label class="form-label">Remarks (optional)</label>
            <textarea id="retRemarks" class="form-control" rows="3" placeholder="Any notes…" disabled></textarea>
          </div>
          <button class="btn btn-warning btn-block" id="returnBtn" onclick="returnDevices()" disabled>
            ↩️ Mark as Returned
          </button>
          <div id="returnAlert"></div>
        </div>
      </div>
    </div>

    <!-- ══════════════════════════════════
         DASHBOARD
    ══════════════════════════════════ -->
    <div id="page-dashboard" class="page">
      <div class="page-header"><h1>Dashboard</h1><p>Live overview of device distribution</p></div>
      <div class="stat-grid" id="dashStats"></div>
      <div class="two-col">
        <div class="card card-pad"><div class="card-title">📊 Department Breakdown</div><div id="deptBreakdown"></div></div>
        <div class="card card-pad"><div class="card-title">⏰ Recent Activity</div><div id="recentActivity"></div></div>
      </div>
    </div>

    <!-- ══════════════════════════════════
         RECORDS
    ══════════════════════════════════ -->
    <div id="page-records" class="page">
      <div class="page-header"><h1>Transaction Records</h1><p>All device issue and return history</p></div>
      <div class="card">
        <div class="card-pad" style="padding-bottom:0">
          <div class="filter-bar">
            <select id="fStatus" class="form-control" onchange="loadRecords()">
              <option value="">All Status</option><option value="issued">Issued</option><option value="returned">Returned</option>
            </select>
            <select id="fDept" class="form-control" onchange="loadRecords()"><option value="">All Departments</option></select>
            <input id="fSearch" class="form-control" placeholder="Search name / reg no…"
              oninput="loadRecords()" style="min-width:200px">
            <span class="count-pill" id="recordCount">— records</span>
          </div>
        </div>
        <div class="table-wrap">
          <table>
            <thead><tr><th>#</th><th>Reg No</th><th>Name</th><th>Dept</th><th>Sec</th>
              <th>iPad</th><th>Stylus</th><th>Issued At</th><th>Returned At</th><th>Condition</th><th>Status</th></tr></thead>
            <tbody id="recordsBody"></tbody>
          </table>
        </div>
      </div>
    </div>

    <!-- ══════════════════════════════════
         STUDENTS — Add + Bulk Upload combined
    ══════════════════════════════════ -->
    <div id="page-students" class="page">
      <div class="page-header">
        <h1>Students</h1>
        <p>Add individual students or bulk import from CSV, then browse the directory</p>
      </div>
      <div class="two-col">

        <!-- LEFT: Combined Add + Upload card -->
        <div class="card card-pad">
          <div class="inner-tabs">
            <div class="inner-tab active" onclick="switchInnerTab('addTab','uploadTab',this)">➕ Add Student</div>
            <div class="inner-tab" onclick="switchInnerTab('uploadTab','addTab',this)">📁 Bulk Upload</div>
          </div>

          <!-- ADD TAB -->
          <div class="inner-tab-pane active" id="addTab">
            <div class="form-group"><label class="form-label">Register Number *</label>
              <input id="sRegNo" class="form-control" placeholder="e.g. CS2024001"></div>
            <div class="form-group"><label class="form-label">Full Name *</label>
              <input id="sName" class="form-control" placeholder="Student full name"></div>
            <div class="form-group"><label class="form-label">Department *</label>
              <input id="sDept" class="form-control" placeholder="e.g. Computer Science"></div>
            <div class="form-group"><label class="form-label">Section *</label>
              <input id="sSec" class="form-control" placeholder="e.g. A"></div>
            <button class="btn btn-primary btn-block" onclick="addStudent()">Add Student</button>
            <div id="studentAddAlert"></div>
          </div>

          <!-- BULK UPLOAD TAB -->
          <div class="inner-tab-pane" id="uploadTab">
            <div class="upload-zone" id="dropZone"
              onclick="document.getElementById('csvFile').click()"
              ondragover="event.preventDefault();this.classList.add('drag')"
              ondragleave="this.classList.remove('drag')"
              ondrop="handleDrop(event)">
              <div class="upload-zone-icon">📂</div>
              <p><strong>Click to browse</strong> or drag & drop</p>
              <p style="font-size:12px;margin-top:4px">Accepted: .csv — Max 5000 rows</p>
            </div>
            <input type="file" id="csvFile" accept=".csv" onchange="handleFileSelect(event)">
            <div class="progress-wrap" id="progressWrap">
              <div style="display:flex;justify-content:space-between;font-size:12px;margin-bottom:4px">
                <span>Uploading…</span><span id="progressPct">0%</span>
              </div>
              <div class="progress-bar-bg"><div class="progress-bar-fill" id="progressBar" style="width:0%"></div></div>
            </div>
            <div id="uploadAlert"></div>
            <button class="btn btn-outline btn-sm" style="margin-top:12px" onclick="downloadTemplate()">⬇️ Download Template</button>

            <div style="margin-top:16px">
              <div class="csv-code">
                <div class="ch">sno,reg_no,name,department,section</div>
                <div class="cr">1,CS2024001,Arjun Mehta,Computer Science,A</div>
                <div class="cr">2,EC2024001,Priya Sharma,Electronics,B</div>
              </div>
              <div class="alert alert-info" style="margin-top:10px;font-size:12px">
                💡 Duplicate reg_no values are skipped automatically.
              </div>
            </div>

            <div class="card" id="previewCard" style="display:none;margin-top:14px;border:1px solid #e5e7eb;box-shadow:none">
              <div class="card-pad" style="padding:14px">
                <div style="font-size:13px;font-weight:700;margin-bottom:10px">👁️ Preview — First 10 rows</div>
                <div class="table-wrap" id="csvPreview"></div>
                <div style="margin-top:10px;display:flex;gap:8px">
                  <button class="btn btn-primary btn-sm" id="confirmUploadBtn" onclick="confirmUpload()">✅ Confirm & Upload</button>
                  <button class="btn btn-outline btn-sm" onclick="cancelUpload()">✖ Cancel</button>
                </div>
              </div>
            </div>
          </div>
        </div>

        <!-- RIGHT: Student directory -->
        <div class="card" style="display:flex;flex-direction:column;min-height:500px">
          <div style="padding:16px 20px;border-bottom:1px solid #f3f4f6;display:flex;justify-content:space-between;align-items:center">
            <span style="font-size:14px;font-weight:700">All Students</span>
            <input id="studentSearch" class="form-control" placeholder="Search…"
              style="width:160px;font-size:12.5px;padding:6px 10px" oninput="loadStudents()">
          </div>
          <div style="overflow-y:auto;flex:1;padding:0 20px" id="studentList"></div>
        </div>
      </div>
    </div>

    <!-- ══════════════════════════════════
         REPORTS
    ══════════════════════════════════ -->
    <div id="page-reports" class="page">
      <div class="page-header"><h1>Reports & Export</h1><p>Download data in Excel or CSV format</p></div>
      <div class="two-col" style="margin-bottom:20px">
        <div class="info-panel">
          <div class="info-tabs"><div class="info-tab active">Database Info</div></div>
          <div class="info-tab-content active">
            <div class="info-section-title">Storage Details</div>
            <div id="storageInfo"></div>
          </div>
        </div>
        <div class="card card-pad">
          <div class="card-title">⚙️ Filter Options</div>
          <div class="form-group"><label class="form-label">Date Range</label>
            <div style="display:flex;gap:8px">
              <input type="date" id="rFrom" class="form-control">
              <input type="date" id="rTo" class="form-control">
            </div>
          </div>
          <div class="form-group"><label class="form-label">Department</label>
            <select id="rDept" class="form-control"><option value="">All Departments</option></select>
          </div>
          <div class="form-group"><label class="form-label">Status</label>
            <select id="rStatus" class="form-control">
              <option value="">All</option><option value="issued">Issued Only</option><option value="returned">Returned Only</option>
            </select>
          </div>
        </div>
      </div>
      <div class="section-divider">📥 Download Reports</div>
      <div class="report-grid">
        <div class="report-btn" onclick="downloadReport('transactions_excel')"><div class="report-btn-icon">📊</div><div><div class="report-btn-title">Full Transaction Report</div><div class="report-btn-sub">Excel (.xlsx) — All records</div></div></div>
        <div class="report-btn" onclick="downloadReport('students_excel')"><div class="report-btn-icon">🎓</div><div><div class="report-btn-title">Student Directory</div><div class="report-btn-sub">Excel (.xlsx) — All students</div></div></div>
        <div class="report-btn" onclick="downloadReport('summary_excel')"><div class="report-btn-icon">📈</div><div><div class="report-btn-title">Summary Report</div><div class="report-btn-sub">Excel (.xlsx) — Dept statistics</div></div></div>
        <div class="report-btn" onclick="downloadReport('transactions_csv')"><div class="report-btn-icon">📋</div><div><div class="report-btn-title">Transactions CSV</div><div class="report-btn-sub">Raw CSV — All transactions</div></div></div>
        <div class="report-btn" onclick="downloadReport('students_csv')"><div class="report-btn-icon">📄</div><div><div class="report-btn-title">Students CSV</div><div class="report-btn-sub">Raw CSV — Student list</div></div></div>
        <div class="report-btn" onclick="downloadReport('issued_csv')"><div class="report-btn-icon">⚠️</div><div><div class="report-btn-title">Currently Issued</div><div class="report-btn-sub">CSV — Pending returns</div></div></div>
      </div>
    </div>

  </div><!-- .main -->
</div><!-- .app-body -->

<script>
/* ── NAVIGATION ── */
function nav(page, el) {
  document.querySelectorAll('.nav-item').forEach(n=>n.classList.remove('active'));
  document.querySelectorAll('.page').forEach(p=>p.classList.remove('active'));
  if(el) el.classList.add('active');
  document.getElementById('page-'+page).classList.add('active');
  if(page==='dashboard') loadDashboard();
  if(page==='records')   { loadRecords(); loadDeptFilter(); }
  if(page==='students')  loadStudents();
  if(page==='reports')   { loadStorageInfo(); loadReportDepts(); }
}

/* ── INNER TABS (Students page) ── */
function switchInnerTab(showId, hideId, el) {
  document.getElementById(showId).classList.add('active');
  document.getElementById(hideId).classList.remove('active');
  el.parentElement.querySelectorAll('.inner-tab').forEach(t=>t.classList.remove('active'));
  el.classList.add('active');
}

/* ── AVATAR HELPERS ── */
const AV_COLORS=['av-green','av-blue','av-purple','av-amber','av-rose','av-cyan','av-indigo','av-teal'];
function avatarClass(name){ return AV_COLORS[(name.charCodeAt(0)||0)%AV_COLORS.length]; }
function initials(name){ return name.split(' ').map(w=>w[0]).join('').toUpperCase().slice(0,2); }
function mkBadge(status){ return status==='issued'
  ?`<span style="display:inline-block;padding:2px 8px;border-radius:20px;font-size:11px;font-weight:700;background:#fef3c7;color:#92400e;border:1px solid #fcd34d">Issued</span>`
  :`<span style="display:inline-block;padding:2px 8px;border-radius:20px;font-size:11px;font-weight:700;background:#dcfce7;color:#166534;border:1px solid #86efac">Returned</span>`; }

/* ══════════════════════════════════════
   ISSUE — live multi-result search
══════════════════════════════════════ */
let currentStudent = null;

async function searchStudentLive(ctx) {
  const q = document.getElementById('issueSearch').value.trim();
  const resEl = document.getElementById('issueResults');
  const cardEl = document.getElementById('issueStudentResult');

  if (!q) { resEl.style.display='none'; return; }

  const data = await (await fetch('/api/students/search-multi?q='+encodeURIComponent(q)+'&mode=issue')).json();

  if (!data.results || data.results.length === 0) {
    resEl.style.display = 'block';
    resEl.innerHTML = `<div class="search-no-result">❌ No student found for "<b>${q}</b>"</div>`;
    currentStudent = null;
    cardEl.innerHTML = '';
    lockIssueForm(true);
    return;
  }

  resEl.style.display = 'block';
  resEl.innerHTML = data.results.map(s => {
    const hasIssue = !!s.active_issue;
    const badge = hasIssue
      ? `<span class="sr-badge issued">Issued</span>`
      : `<span class="sr-badge available">Available</span>`;
    return `<div class="search-result-item" onclick="selectIssueStudent(${JSON.stringify(s).replace(/"/g,'&quot;')})">
      <div class="sr-avatar ${avatarClass(s.name)}">${initials(s.name)}</div>
      <div style="flex:1;min-width:0">
        <div class="sr-name">${s.name}</div>
        <div class="sr-meta">${s.reg_no} · ${s.department} · Sec ${s.section}</div>
      </div>
      ${badge}
    </div>`;
  }).join('');
}

function selectIssueStudent(s) {
  currentStudent = s;
  const resEl = document.getElementById('issueResults');
  const cardEl = document.getElementById('issueStudentResult');

  // Highlight selected row
  resEl.querySelectorAll('.search-result-item').forEach(el => el.classList.remove('selected'));
  event.currentTarget.classList.add('selected');

  const hasIssue = !!s.active_issue;
  const warn = hasIssue
    ? `<div class="alert alert-warning" style="margin-top:8px">⚠️ Already has active issue — iPad: ${s.active_issue.ipad_no}, Stylus: ${s.active_issue.stylus_no}</div>`
    : '';

  cardEl.innerHTML = `
    <div class="found-card">
      <button class="found-clear" onclick="clearIssueSelection()" title="Clear">✕</button>
      <div style="display:flex;justify-content:space-between;align-items:center">
        <span class="found-reg">${s.reg_no}</span>
        <span style="font-size:11px;color:#9ca3af">S.No: ${s.sno}</span>
      </div>
      <div class="found-name">${s.name}</div>
      <div class="found-meta">
        <span>🏛️ ${s.department}</span>
        <span>📌 Section ${s.section}</span>
      </div>
    </div>${warn}`;

  lockIssueForm(hasIssue);
}

function clearIssueSelection() {
  currentStudent = null;
  document.getElementById('issueStudentResult').innerHTML = '';
  document.getElementById('issueResults').style.display = 'none';
  document.getElementById('issueSearch').value = '';
  lockIssueForm(true);
}

function lockIssueForm(disabled) {
  ['ipadNo','stylusNo','issuedBy','issueBtn'].forEach(id=>document.getElementById(id).disabled=disabled);
}

async function issueDevices() {
  const ipad   = document.getElementById('ipadNo').value.trim();
  const stylus = document.getElementById('stylusNo').value.trim();
  const by     = document.getElementById('issuedBy').value.trim() || 'Staff';
  if (!ipad || !stylus) { showAlert('issueAlert','Enter both iPad and Stylus numbers.','warning'); return; }
  const d = await (await fetch('/api/issue',{method:'POST',headers:{'Content-Type':'application/json'},
    body:JSON.stringify({reg_no:currentStudent.reg_no,ipad_no:ipad,stylus_no:stylus,issued_by:by})})).json();
  if (d.success) {
    showAlert('issueAlert',`✅ Issued to ${currentStudent.name} — iPad: ${ipad}, Stylus: ${stylus}`,'success');
    clearIssueSelection();
    ['ipadNo','stylusNo'].forEach(id=>document.getElementById(id).value='');
  } else { showAlert('issueAlert','❌ '+d.error,'danger'); }
}

/* ══════════════════════════════════════
   RETURN — live multi-result search
══════════════════════════════════════ */
let currentIssue = null;

async function searchReturnLive() {
  const q = document.getElementById('returnSearch').value.trim();
  const resEl = document.getElementById('returnResults');
  const cardEl = document.getElementById('returnStudentResult');

  if (!q) { resEl.style.display='none'; return; }

  const data = await (await fetch('/api/students/search-multi?q='+encodeURIComponent(q)+'&mode=return')).json();

  if (!data.results || data.results.length === 0) {
    resEl.style.display = 'block';
    resEl.innerHTML = `<div class="search-no-result">❌ No active issue found for "<b>${q}</b>"</div>`;
    currentIssue = null;
    cardEl.innerHTML = '';
    lockReturnForm(true);
    return;
  }

  resEl.style.display = 'block';
  resEl.innerHTML = data.results.map(r => `
    <div class="search-result-item" onclick="selectReturnRecord(${JSON.stringify(r).replace(/"/g,'&quot;')})">
      <div class="sr-avatar ${avatarClass(r.name)}">${initials(r.name)}</div>
      <div style="flex:1;min-width:0">
        <div class="sr-name">${r.name}</div>
        <div class="sr-meta">${r.reg_no} · iPad: ${r.ipad_no} · Stylus: ${r.stylus_no}</div>
      </div>
      <span class="sr-badge issued">Issued</span>
    </div>`).join('');
}

function selectReturnRecord(r) {
  currentIssue = r;
  const resEl = document.getElementById('returnResults');
  const cardEl = document.getElementById('returnStudentResult');

  resEl.querySelectorAll('.search-result-item').forEach(el=>el.classList.remove('selected'));
  event.currentTarget.classList.add('selected');

  cardEl.innerHTML = `
    <div class="found-card">
      <button class="found-clear" onclick="clearReturnSelection()" title="Clear">✕</button>
      <div style="display:flex;justify-content:space-between;align-items:center">
        <span class="found-reg">${r.reg_no}</span>
        ${mkBadge('issued')}
      </div>
      <div class="found-name">${r.name}</div>
      <div class="found-meta">
        <span>🏛️ ${r.department}</span>
        <span>📌 Sec ${r.section}</span>
      </div>
      <div style="margin-top:8px;font-size:12.5px;color:#374151;display:flex;gap:16px">
        <span>📱 iPad: <b>${r.ipad_no}</b></span>
        <span>✏️ Stylus: <b>${r.stylus_no}</b></span>
      </div>
      <div style="font-size:11.5px;color:#9ca3af;margin-top:4px">Issued: ${r.issued_at}</div>
    </div>`;
  lockReturnForm(false);
}

function clearReturnSelection() {
  currentIssue = null;
  document.getElementById('returnStudentResult').innerHTML = '';
  document.getElementById('returnResults').style.display = 'none';
  document.getElementById('returnSearch').value = '';
  lockReturnForm(true);
}

function lockReturnForm(disabled) {
  ['retCondition','retRemarks','returnBtn'].forEach(id=>document.getElementById(id).disabled=disabled);
}

async function returnDevices() {
  const d = await (await fetch('/api/return',{method:'POST',headers:{'Content-Type':'application/json'},
    body:JSON.stringify({transaction_id:currentIssue.id,
      condition:document.getElementById('retCondition').value,
      remarks:document.getElementById('retRemarks').value})})).json();
  if (d.success) {
    showAlert('returnAlert',`✅ Return processed for ${currentIssue.name} — iPad: ${currentIssue.ipad_no}`,'success');
    clearReturnSelection();
    document.getElementById('retRemarks').value='';
  } else { showAlert('returnAlert','❌ '+d.error,'danger'); }
}

/* ── DASHBOARD ── */
async function loadDashboard() {
  const d = await (await fetch('/api/dashboard')).json();
  document.getElementById('dashStats').innerHTML=[
    {icon:'📋',n:d.total,   l:'Total Transactions',cls:'blue'},
    {icon:'📤',n:d.issued,  l:'Currently Issued',  cls:'amber'},
    {icon:'📥',n:d.returned,l:'Returned',           cls:'green'},
    {icon:'🎓',n:d.students,l:'Total Students',     cls:'purple'},
  ].map(s=>`<div class="stat-card"><div class="stat-icon ${s.cls}">${s.icon}</div>
    <div><div class="stat-label">${s.l}</div><div class="stat-value">${s.n}</div></div></div>`).join('');

  document.getElementById('deptBreakdown').innerHTML = d.by_dept.length ?
    d.by_dept.map(r=>`<div style="margin-bottom:12px">
      <div style="display:flex;justify-content:space-between;font-size:12.5px;margin-bottom:4px">
        <span style="font-weight:600;color:#374151">${r.department}</span>
        <span style="color:#9ca3af">${r.issued} out · ${r.returned} returned</span>
      </div>
      <div style="background:#f3f4f6;border-radius:4px;height:8px;overflow:hidden">
        <div style="height:8px;background:#22c55e;border-radius:4px;width:${r.total?Math.round(r.returned/r.total*100):0}%"></div>
      </div></div>`).join('') : '<p style="color:#9ca3af;font-size:13px">No data yet.</p>';

  document.getElementById('recentActivity').innerHTML = d.recent.length ?
    d.recent.map(r=>`<div class="member-item">
      <div class="member-avatar ${avatarClass(r.name)}">${initials(r.name)}</div>
      <div><div class="member-name">${r.name}</div><div class="member-email">${r.reg_no} · ${r.ipad_no||'—'}</div></div>
      <div class="member-status">${mkBadge(r.status)}</div>
    </div>`).join('') : '<p style="color:#9ca3af;font-size:13px">No activity yet.</p>';
}

/* ── RECORDS ── */
async function loadDeptFilter() {
  const d = await (await fetch('/api/departments')).json();
  document.getElementById('fDept').innerHTML='<option value="">All Departments</option>'+d.map(x=>`<option>${x}</option>`).join('');
}
async function loadRecords() {
  const status=document.getElementById('fStatus').value;
  const dept=document.getElementById('fDept').value;
  const q=document.getElementById('fSearch').value;
  const d=await(await fetch(`/api/records?status=${status}&dept=${encodeURIComponent(dept)}&q=${encodeURIComponent(q)}`)).json();
  document.getElementById('recordCount').textContent=d.length+' records';
  document.getElementById('recordsBody').innerHTML=d.map((r,i)=>`<tr>
    <td style="color:#9ca3af">${i+1}</td>
    <td><code>${r.reg_no}</code></td>
    <td style="font-weight:600">${r.name}</td>
    <td>${r.department}</td><td>${r.section}</td>
    <td><code>${r.ipad_no}</code></td><td><code>${r.stylus_no}</code></td>
    <td style="color:#9ca3af;font-size:12px">${r.issued_at||''}</td>
    <td style="color:#9ca3af;font-size:12px">${r.returned_at||'—'}</td>
    <td style="font-size:12px">${r.condition||'—'}</td>
    <td>${mkBadge(r.status)}</td>
  </tr>`).join('')||'<tr><td colspan="11" style="text-align:center;padding:32px;color:#9ca3af">No records found.</td></tr>';
}

/* ── STUDENTS ── */
async function loadStudents() {
  const q = document.getElementById('studentSearch').value;
  const d = await (await fetch('/api/students?q='+encodeURIComponent(q))).json();
  document.getElementById('studentList').innerHTML = d.map(s=>`<div class="member-item">
    <div class="member-avatar ${avatarClass(s.name)}">${initials(s.name)}</div>
    <div>
      <div class="member-name">${s.name}</div>
      <div class="member-email">${s.reg_no} · ${s.department} · Sec ${s.section}</div>
    </div>
    <div class="member-status" style="font-size:11px;color:#9ca3af">#${s.sno}</div>
  </div>`).join('')||'<p style="padding:24px;text-align:center;color:#9ca3af;font-size:13px">No students found.</p>';
}

async function addStudent() {
  const body={reg_no:document.getElementById('sRegNo').value.trim(),
    name:document.getElementById('sName').value.trim(),
    department:document.getElementById('sDept').value.trim(),
    section:document.getElementById('sSec').value.trim()};
  if(!body.reg_no||!body.name||!body.department||!body.section){
    showAlert('studentAddAlert','All fields are required.','warning'); return; }
  const d=await(await fetch('/api/student/add',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(body)})).json();
  if(d.success){
    showAlert('studentAddAlert','✅ Student '+body.name+' added!','success');
    ['sRegNo','sName','sDept','sSec'].forEach(id=>document.getElementById(id).value='');
    loadStudents();
  } else { showAlert('studentAddAlert','❌ '+d.error,'danger'); }
}

/* ── BULK UPLOAD ── */
let csvRows=[];
function handleDrop(e){e.preventDefault();document.getElementById('dropZone').classList.remove('drag');if(e.dataTransfer.files[0])processCSVFile(e.dataTransfer.files[0]);}
function handleFileSelect(e){if(e.target.files[0])processCSVFile(e.target.files[0]);}
function processCSVFile(file){
  if(!file.name.endsWith('.csv')){showAlert('uploadAlert','Only .csv files allowed.','danger');return;}
  const reader=new FileReader();
  reader.onload=e=>{
    const lines=e.target.result.trim().split('\n');
    const headers=lines[0].toLowerCase().split(',').map(h=>h.trim());
    const missing=['sno','reg_no','name','department','section'].filter(h=>!headers.includes(h));
    if(missing.length){showAlert('uploadAlert','Missing columns: '+missing.join(', '),'danger');return;}
    csvRows=lines.slice(1).map(l=>l.split(',').map(x=>x.trim())).filter(r=>r.length>=5);
    document.getElementById('csvPreview').innerHTML=`<table><thead><tr>${headers.map(h=>`<th>${h}</th>`).join('')}</tr></thead>
      <tbody>${csvRows.slice(0,10).map(r=>`<tr>${r.map(c=>`<td>${c}</td>`).join('')}</tr>`).join('')}</tbody></table>`;
    document.getElementById('previewCard').style.display='block';
    showAlert('uploadAlert',`📋 ${csvRows.length} rows ready. Review below and confirm.`,'info');
  };
  reader.readAsText(file);
}
async function confirmUpload(){
  const btn=document.getElementById('confirmUploadBtn');btn.disabled=true;btn.textContent='Uploading…';
  document.getElementById('progressWrap').style.display='block';
  document.getElementById('progressBar').style.width='40%';document.getElementById('progressPct').textContent='40%';
  const d=await(await fetch('/api/students/bulk',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({rows:csvRows})})).json();
  document.getElementById('progressBar').style.width='100%';document.getElementById('progressPct').textContent='100%';
  showAlert('uploadAlert',`✅ Upload complete — ${d.inserted} added, ${d.skipped} skipped (duplicates).`,'success');
  document.getElementById('previewCard').style.display='none';
  btn.disabled=false;btn.textContent='✅ Confirm & Upload';
  csvRows=[];document.getElementById('csvFile').value='';
  loadStudents();
}
function cancelUpload(){
  csvRows=[];document.getElementById('previewCard').style.display='none';
  document.getElementById('csvFile').value='';
  document.getElementById('progressWrap').style.display='none';
  document.getElementById('uploadAlert').innerHTML='';
}
function downloadTemplate(){
  const csv='sno,reg_no,name,department,section\n1,CS2024001,Student Name,Computer Science,A\n2,EC2024001,Student Name,Electronics,B';
  const a=document.createElement('a');a.href='data:text/csv;charset=utf-8,'+encodeURIComponent(csv);a.download='student_template.csv';a.click();
}

/* ── REPORTS ── */
async function loadStorageInfo(){
  const d=await(await fetch('/api/storage_info')).json();
  document.getElementById('storageInfo').innerHTML=`
    <div class="info-grid" style="grid-template-columns:1fr 1fr">
      <div class="info-field"><span class="info-field-icon">🐘</span><div><div class="info-field-label">Database Engine</div><div class="info-field-value">PostgreSQL</div></div></div>
      <div class="info-field"><span class="info-field-icon">🌐</span><div><div class="info-field-label">Host</div><div class="info-field-value">${d.host}:${d.port}</div></div></div>
      <div class="info-field"><span class="info-field-icon">🗄️</span><div><div class="info-field-label">Database</div><div class="info-field-value">${d.dbname}</div></div></div>
      <div class="info-field"><span class="info-field-icon">💾</span><div><div class="info-field-label">DB Size</div><div class="info-field-value">${d.size_mb} MB</div></div></div>
      <div class="info-field"><span class="info-field-icon">🎓</span><div><div class="info-field-label">Students</div><div class="info-field-value">${d.students} records</div></div></div>
      <div class="info-field"><span class="info-field-icon">📋</span><div><div class="info-field-label">Transactions</div><div class="info-field-value">${d.transactions} records</div></div></div>
    </div>
    <div class="alert alert-success" style="font-size:12px;margin-top:4px">✅ Data stored in PostgreSQL. Use pg_dump for backups.</div>`;
}
async function loadReportDepts(){
  const d=await(await fetch('/api/departments')).json();
  document.getElementById('rDept').innerHTML='<option value="">All Departments</option>'+d.map(x=>`<option>${x}</option>`).join('');
}
function downloadReport(type){
  const from=document.getElementById('rFrom')?.value||'';
  const to=document.getElementById('rTo')?.value||'';
  const dept=document.getElementById('rDept')?.value||'';
  const status=document.getElementById('rStatus')?.value||'';
  window.location.href=`/api/report/${type}?from=${from}&to=${to}&dept=${encodeURIComponent(dept)}&status=${status}`;
}

/* ── UTIL ── */
function showAlert(id,msg,type){
  document.getElementById(id).innerHTML=`<div class="alert alert-${type}">${msg}</div>`;
  setTimeout(()=>{const el=document.getElementById(id);if(el)el.innerHTML='';},5000);
}
</script>
</body>
</html>"""


# ── API ROUTES ────────────────────────────────────────────

@app.route("/")
def index():
    return render_template_string(HTML)

@app.route("/api/stats")
def api_stats():
    return jsonify({
        "issued":   scalar("SELECT COUNT(*) FROM transactions WHERE status='issued'"),
        "returned": scalar("SELECT COUNT(*) FROM transactions WHERE status='returned'"),
        "students": scalar("SELECT COUNT(*) FROM students"),
    })

# Multi-result search — used by Issue page (mode=issue) and Return page (mode=return)
@app.route("/api/students/search-multi")
def search_multi():
    q    = request.args.get("q", "").strip()
    mode = request.args.get("mode", "issue")   # 'issue' or 'return'
    if not q:
        return jsonify({"results": []})

    if mode == "return":
        # Only students with an active issued transaction
        rows = query("""
            SELECT * FROM transactions
            WHERE status = 'issued'
              AND (reg_no ILIKE %s OR name ILIKE %s)
            ORDER BY issued_at DESC LIMIT 20
        """, (f"%{q}%", f"%{q}%"))
        for r in rows:
            ts(r, "issued_at", "returned_at")
        return jsonify({"results": rows})
    else:
        # All matching students, annotated with active_issue
        students = query("""
            SELECT * FROM students
            WHERE reg_no ILIKE %s OR name ILIKE %s
            ORDER BY name LIMIT 20
        """, (f"%{q}%", f"%{q}%"))
        for s in students:
            ts(s, "created_at")
            active = query(
                "SELECT * FROM transactions WHERE reg_no = %s AND status = 'issued' LIMIT 1",
                (s["reg_no"],), fetch="one"
            )
            if active:
                ts(active, "issued_at", "returned_at")
            s["active_issue"] = active
        return jsonify({"results": students})

@app.route("/api/issue", methods=["POST"])
def issue_device():
    data = request.json
    student = query("SELECT * FROM students WHERE reg_no = %s", (data["reg_no"],), fetch="one")
    if not student:
        return jsonify({"success": False, "error": "Student not found"})
    existing = query(
        "SELECT id FROM transactions WHERE reg_no = %s AND status = 'issued'",
        (data["reg_no"],), fetch="one"
    )
    if existing:
        return jsonify({"success": False, "error": "Student already has active issue"})
    query("""
        INSERT INTO transactions (reg_no, name, department, section, ipad_no, stylus_no, issued_by)
        VALUES (%s, %s, %s, %s, %s, %s, %s)
    """, (student["reg_no"], student["name"], student["department"], student["section"],
          data["ipad_no"], data["stylus_no"], data.get("issued_by", "Staff")), fetch="none")
    return jsonify({"success": True})

@app.route("/api/return", methods=["POST"])
def return_device():
    data = request.json
    query("""
        UPDATE transactions
        SET status = 'returned', returned_at = NOW(), condition = %s, remarks = %s
        WHERE id = %s AND status = 'issued'
    """, (data.get("condition"), data.get("remarks"), data["transaction_id"]), fetch="none")
    return jsonify({"success": True})

@app.route("/api/dashboard")
def dashboard():
    by_dept = query("""
        SELECT department,
               SUM(CASE WHEN status='issued'   THEN 1 ELSE 0 END) AS issued,
               SUM(CASE WHEN status='returned' THEN 1 ELSE 0 END) AS returned,
               COUNT(*) AS total
        FROM transactions GROUP BY department ORDER BY department
    """)
    recent = query("SELECT * FROM transactions ORDER BY id DESC LIMIT 6")
    for r in recent:
        ts(r, "issued_at", "returned_at")
    return jsonify({
        "total":    scalar("SELECT COUNT(*) FROM transactions"),
        "issued":   scalar("SELECT COUNT(*) FROM transactions WHERE status='issued'"),
        "returned": scalar("SELECT COUNT(*) FROM transactions WHERE status='returned'"),
        "students": scalar("SELECT COUNT(*) FROM students"),
        "by_dept":  by_dept,
        "recent":   recent,
    })

@app.route("/api/records")
def get_records():
    status = request.args.get("status", "")
    dept   = request.args.get("dept",   "")
    q      = request.args.get("q",      "")
    sql    = "SELECT * FROM transactions WHERE 1=1"
    params = []
    if status: sql += " AND status = %s";               params.append(status)
    if dept:   sql += " AND department = %s";           params.append(dept)
    if q:      sql += " AND (reg_no ILIKE %s OR name ILIKE %s)"; params += [f"%{q}%", f"%{q}%"]
    sql += " ORDER BY id DESC"
    rows = query(sql, params)
    for r in rows:
        ts(r, "issued_at", "returned_at")
    return jsonify(rows)

@app.route("/api/students")
def get_students():
    q = request.args.get("q", "")
    rows = query(
        "SELECT * FROM students WHERE name ILIKE %s OR reg_no ILIKE %s ORDER BY sno",
        (f"%{q}%", f"%{q}%")
    )
    for r in rows:
        ts(r, "created_at")
    return jsonify(rows)

@app.route("/api/student/add", methods=["POST"])
def add_student():
    d = request.json
    try:
        query("""
            INSERT INTO students (reg_no, name, department, section)
            VALUES (%s, %s, %s, %s)
        """, (d["reg_no"], d["name"], d["department"], d["section"]), fetch="none")
        return jsonify({"success": True})
    except pg_errors.UniqueViolation:
        return jsonify({"success": False, "error": "Register number already exists"})

@app.route("/api/students/bulk", methods=["POST"])
def bulk_upload():
    rows = request.json.get("rows", [])
    inserted = skipped = 0
    for row in rows:
        if len(row) < 5:
            continue
        try:
            query("""
                INSERT INTO students (reg_no, name, department, section)
                VALUES (%s, %s, %s, %s)
            """, (row[1].strip(), row[2].strip(), row[3].strip(), row[4].strip()), fetch="none")
            inserted += 1
        except pg_errors.UniqueViolation:
            skipped += 1
    return jsonify({"inserted": inserted, "skipped": skipped})

@app.route("/api/departments")
def get_departments():
    rows = query("SELECT DISTINCT department FROM students ORDER BY department")
    return jsonify([r["department"] for r in rows])

@app.route("/api/storage_info")
def storage_info():
    size_mb = scalar("SELECT ROUND(pg_database_size(current_database()) / 1024.0 / 1024.0, 2)")
    return jsonify({
        "host":         DB_CONFIG["host"],
        "port":         DB_CONFIG["port"],
        "dbname":       DB_CONFIG["dbname"],
        "size_mb":      size_mb,
        "students":     scalar("SELECT COUNT(*) FROM students"),
        "transactions": scalar("SELECT COUNT(*) FROM transactions"),
        "issued":       scalar("SELECT COUNT(*) FROM transactions WHERE status='issued'"),
    })

# ── REPORT HELPERS ────────────────────────────────────────

def get_filtered_txn(from_date, to_date, dept, status):
    sql    = "SELECT * FROM transactions WHERE 1=1"
    params = []
    if from_date: sql += " AND issued_at::date >= %s"; params.append(from_date)
    if to_date:   sql += " AND issued_at::date <= %s"; params.append(to_date)
    if dept:      sql += " AND department = %s";       params.append(dept)
    if status:    sql += " AND status = %s";           params.append(status)
    sql += " ORDER BY id DESC"
    rows = query(sql, params)
    for r in rows:
        ts(r, "issued_at", "returned_at")
    return rows

def style_header(ws, headers, color="16A34A"):
    fill   = PatternFill("solid", fgColor=color)
    font   = Font(bold=True, color="FFFFFF", size=11)
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'),  bottom=Side(style='thin'))
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = fill; cell.font = font
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = max(len(h) + 4, 14)

@app.route("/api/report/<rtype>")
def download_report(rtype):
    fd     = request.args.get("from",   "")
    td     = request.args.get("to",     "")
    dept   = request.args.get("dept",   "")
    status = request.args.get("status", "")

    if rtype == "transactions_csv":
        rows = get_filtered_txn(fd, td, dept, status)
        out  = io.StringIO(); w = csv.writer(out)
        w.writerow(["S.No","Reg No","Name","Department","Section","iPad No","Stylus No",
                    "Issued At","Returned At","Status","Condition","Remarks","Issued By"])
        for i, r in enumerate(rows, 1):
            w.writerow([i, r["reg_no"], r["name"], r["department"], r["section"],
                        r["ipad_no"], r["stylus_no"], r["issued_at"], r["returned_at"] or "",
                        r["status"], r["condition"] or "", r["remarks"] or "", r["issued_by"] or ""])
        out.seek(0)
        return send_file(io.BytesIO(out.getvalue().encode()), download_name="transactions.csv",
                         as_attachment=True, mimetype="text/csv")

    elif rtype == "students_csv":
        rows = query("SELECT * FROM students ORDER BY sno")
        for r in rows: ts(r, "created_at")
        out  = io.StringIO(); w = csv.writer(out)
        w.writerow(["S.No","Reg No","Name","Department","Section","Added At"])
        for r in rows:
            w.writerow([r["sno"], r["reg_no"], r["name"], r["department"], r["section"], r["created_at"]])
        out.seek(0)
        return send_file(io.BytesIO(out.getvalue().encode()), download_name="students.csv",
                         as_attachment=True, mimetype="text/csv")

    elif rtype == "issued_csv":
        rows = query("SELECT * FROM transactions WHERE status='issued' ORDER BY issued_at")
        for r in rows: ts(r, "issued_at", "returned_at")
        out  = io.StringIO(); w = csv.writer(out)
        w.writerow(["S.No","Reg No","Name","Department","Section","iPad No","Stylus No","Issued At","Issued By"])
        for i, r in enumerate(rows, 1):
            w.writerow([i, r["reg_no"], r["name"], r["department"], r["section"],
                        r["ipad_no"], r["stylus_no"], r["issued_at"], r["issued_by"] or ""])
        out.seek(0)
        return send_file(io.BytesIO(out.getvalue().encode()), download_name="currently_issued.csv",
                         as_attachment=True, mimetype="text/csv")

    elif rtype == "transactions_excel":
        rows = get_filtered_txn(fd, td, dept, status)
        wb   = openpyxl.Workbook(); ws = wb.active; ws.title = "Transactions"
        style_header(ws, ["S.No","Reg No","Name","Department","Section","iPad No","Stylus No",
                           "Issued At","Returned At","Status","Condition","Remarks","Issued By"])
        alt = PatternFill("solid", fgColor="F0FDF4")
        for i, r in enumerate(rows, 1):
            row_data = [i, r["reg_no"], r["name"], r["department"], r["section"],
                        r["ipad_no"], r["stylus_no"], r["issued_at"], r["returned_at"] or "",
                        r["status"], r["condition"] or "", r["remarks"] or "", r["issued_by"] or ""]
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=i+1, column=col, value=val)
                if i % 2 == 0: cell.fill = alt
        out = io.BytesIO(); wb.save(out); out.seek(0)
        return send_file(out, download_name="transactions_report.xlsx", as_attachment=True,
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    elif rtype == "students_excel":
        rows = query("SELECT * FROM students ORDER BY sno")
        for r in rows: ts(r, "created_at")
        wb   = openpyxl.Workbook(); ws = wb.active; ws.title = "Students"
        style_header(ws, ["S.No","Reg No","Name","Department","Section","Added At"], "0F172A")
        alt = PatternFill("solid", fgColor="F9FAFB")
        for i, r in enumerate(rows, 1):
            row_data = [r["sno"], r["reg_no"], r["name"], r["department"], r["section"], r["created_at"]]
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=i+1, column=col, value=val)
                if i % 2 == 0: cell.fill = alt
        out = io.BytesIO(); wb.save(out); out.seek(0)
        return send_file(out, download_name="students_directory.xlsx", as_attachment=True,
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    elif rtype == "summary_excel":
        by_dept = query("""
            SELECT department,
                   COUNT(*) AS total_transactions,
                   SUM(CASE WHEN status='issued'   THEN 1 ELSE 0 END) AS currently_issued,
                   SUM(CASE WHEN status='returned' THEN 1 ELSE 0 END) AS returned,
                   COUNT(DISTINCT reg_no) AS unique_students
            FROM transactions GROUP BY department ORDER BY department
        """)
        by_sec = query("""
            SELECT department, section, COUNT(*) AS total
            FROM transactions GROUP BY department, section ORDER BY department, section
        """)
        wb  = openpyxl.Workbook(); ws = wb.active; ws.title = "Dept Summary"
        style_header(ws, ["Department","Total Transactions","Currently Issued","Returned","Unique Students"])
        alt = PatternFill("solid", fgColor="F0FDF4")
        for i, r in enumerate(by_dept, 1):
            row_data = [r["department"], r["total_transactions"], r["currently_issued"],
                        r["returned"], r["unique_students"]]
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=i+1, column=col, value=val)
                if i % 2 == 0: cell.fill = alt
        ws2 = wb.create_sheet("Section Summary")
        style_header(ws2, ["Department","Section","Transactions"], "0F172A")
        for i, r in enumerate(by_sec, 1):
            ws2.cell(row=i+1, column=1, value=r["department"])
            ws2.cell(row=i+1, column=2, value=r["section"])
            ws2.cell(row=i+1, column=3, value=r["total"])
        out = io.BytesIO(); wb.save(out); out.seek(0)
        return send_file(out, download_name="summary_report.xlsx", as_attachment=True,
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    return jsonify({"error": "Unknown report type"}), 400

if __name__ == "__main__":
    print("=" * 55)
    print("  📱 iPad & Stylus Distribution System (PostgreSQL)")
    print(f"  🌐 Open : http://127.0.0.1:5000")
    print(f"  🐘 DB   : {DB_CONFIG['host']}:{DB_CONFIG['port']}/{DB_CONFIG['dbname']}")
    print("=" * 55)
    app.run(host="0.0.0.0", debug=True, port=5000)
